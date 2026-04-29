"""
Convert kilder.xlsx → document_store.json entries.
Produces the same dict-of-lists format as the existing document_store.json.
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

XLSX_PATH    = Path(r"C:\Users\ChristianGoulignac\Downloads\kilder.xlsx")
OUT_PATH     = Path("./utils/create_lab_vectorindex/document_store_kilder.json")
INDEX_NAME   = "kilder"
FILNAVN_PREFIX = "data\\digiung_lab\\"  # prepended to every filnavn read from the spreadsheet

# Column header (Norwegian) → JSON field
COLMAP = {
    "Tittel":             "tittel",
    "Filnavn":            "filnavn",
    "Publisert årstall":  "publisert_arstall",
    "Publisert år":       "publisert_arstall",
    "Publisert av":       "publisert_av",
    "Type kilde":         "type_kilde",
    "Målgruppe":          "malgruppe",
    "Antall deltakere":   "antall_deltakere",
    "Segment":            "segment",
    "Oppsummering":       "oppsummering",
}

def normalise(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v

def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Empty workbook")

    headers = [(h or "").strip() for h in rows[0]]
    field_idx = {COLMAP[h]: i for i, h in enumerate(headers) if h in COLMAP}
    missing_fields = [f for f in ("tittel", "filnavn") if f not in field_idx]
    if missing_fields:
        sys.exit(f"Missing required column(s) in spreadsheet: {missing_fields}")

    entries = []
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {}
        for field, idx in field_idx.items():
            rec[field] = normalise(row[idx])
        # Coerce publisert_arstall to int when possible
        ar = rec.get("publisert_arstall")
        if isinstance(ar, str) and ar.isdigit():
            rec["publisert_arstall"] = int(ar)
        # antall_deltakere is stored as string in the existing file
        ad = rec.get("antall_deltakere")
        if ad is not None and not isinstance(ad, str):
            rec["antall_deltakere"] = str(ad)
        # Skip rows without title or filename
        if not rec.get("tittel") or not rec.get("filnavn"):
            continue
        # Prepend storage path prefix to filnavn
        rec["filnavn"] = FILNAVN_PREFIX + rec["filnavn"]
        # Ensure all expected keys exist (matching existing file shape)
        for k in ("tittel", "publisert_arstall", "publisert_av", "type_kilde",
                  "malgruppe", "antall_deltakere", "segment", "oppsummering", "filnavn"):
            rec.setdefault(k, None)
        entries.append(rec)

    out = {INDEX_NAME: entries}
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"Wrote {len(entries)} entries under index '{INDEX_NAME}' to {OUT_PATH}")

if __name__ == "__main__":
    main()
